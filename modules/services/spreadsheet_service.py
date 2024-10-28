# services/spreadsheet_service.py

import logging
import threading
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Any, Dict, List, Optional, Union
import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from modules.utilities.logging_manager import setup_logging
from modules.utilities.config_loader import ConfigLoader
from modules.security.encryption_manager import EncryptionManager
from modules.security.authentication import AuthenticationManager


class SpreadsheetServiceError(Exception):
    """Custom exception for SpreadsheetService-related errors."""
    pass


class SpreadsheetService:
    """
    Provides spreadsheet management capabilities, including creating, reading, updating, and formatting
    Excel files. Utilizes pandas and openpyxl libraries to ensure comprehensive spreadsheet handling.
    Supports advanced features like data analysis, chart creation, and conditional formatting.
    Ensures secure handling of files and configurations.
    """

    def __init__(self):
        """
        Initializes the SpreadsheetService with necessary configurations and authentication.
        """
        self.logger = setup_logging('SpreadsheetService')
        self.config_loader = ConfigLoader()
        self.encryption_manager = EncryptionManager()
        self.auth_manager = AuthenticationManager()
        self.lock = threading.Lock()
        self.logger.info("SpreadsheetService initialized successfully.")

    def create_spreadsheet(self, output_path: str, data: Dict[str, List[Any]], sheet_names: Optional[List[str]] = None) -> bool:
        """
        Creates a new Excel spreadsheet with specified data and sheets.

        Args:
            output_path (str): The file path for the created spreadsheet.
            data (Dict[str, List[Any]]): A dictionary where keys are sheet names and values are lists of rows (lists).
            sheet_names (Optional[List[str]], optional): A list of sheet names. If None, keys from data are used. Defaults to None.

        Returns:
            bool: True if the spreadsheet is created successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Creating spreadsheet at '{output_path}' with sheets: {sheet_names if sheet_names else list(data.keys())}.")
            with self.lock:
                wb = Workbook()
                default_sheet = wb.active
                default_sheet.title = sheet_names[0] if sheet_names else list(data.keys())[0]
                for sheet_name, rows in data.items():
                    if sheet_names and sheet_name != sheet_names[0]:
                        ws = wb.create_sheet(title=sheet_name)
                    else:
                        ws = default_sheet
                    for row in rows:
                        ws.append(row)
                        self.logger.debug(f"Added row {row} to sheet '{ws.title}'.")
                wb.save(output_path)
            self.logger.info(f"Spreadsheet created successfully at '{output_path}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error creating spreadsheet at '{output_path}': {e}", exc_info=True)
            return False

    def read_spreadsheet(self, file_path: str, sheet_name: Optional[str] = None) -> Optional[Dict[str, Any]]:
        """
        Reads data from an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (Optional[str], optional): The sheet name to read. If None, all sheets are read. Defaults to None.

        Returns:
            Optional[Dict[str, Any]]: A dictionary where keys are sheet names and values are DataFrames, or None if reading fails.
        """
        try:
            self.logger.debug(f"Reading spreadsheet from '{file_path}' with sheet '{sheet_name}'.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return None
            with self.lock:
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                    self.logger.debug(f"Read data from sheet '{sheet_name}'.")
                    return {sheet_name: df}
                else:
                    dfs = pd.read_excel(file_path, sheet_name=None)
                    self.logger.debug(f"Read data from all sheets.")
                    return dfs
        except Exception as e:
            self.logger.error(f"Error reading spreadsheet '{file_path}': {e}", exc_info=True)
            return None

    def update_spreadsheet(self, file_path: str, sheet_name: str, data: List[List[Any]], append: bool = False) -> bool:
        """
        Updates an existing Excel spreadsheet by adding new data to a specified sheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to update.
            data (List[List[Any]]): A list of rows (lists) to add to the sheet.
            append (bool, optional): Whether to append data to the end of the sheet. Defaults to False.

        Returns:
            bool: True if the spreadsheet is updated successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Updating spreadsheet '{file_path}' in sheet '{sheet_name}' with data: {data}. Append={append}.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False
            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]
                for row in data:
                    if append:
                        ws.append(row)
                        self.logger.debug(f"Appended row {row} to sheet '{sheet_name}'.")
                    else:
                        ws.append(row)
                        self.logger.debug(f"Added row {row} to sheet '{sheet_name}'.")
                wb.save(file_path)
            self.logger.info(f"Spreadsheet '{file_path}' updated successfully in sheet '{sheet_name}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error updating spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def format_spreadsheet(self, file_path: str, sheet_name: str, formatting_rules: Dict[str, Any]) -> bool:
        """
        Applies formatting rules to a specified sheet in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to format.
            formatting_rules (Dict[str, Any]): A dictionary containing formatting rules (e.g., font size, alignment).

        Returns:
            bool: True if the spreadsheet is formatted successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Applying formatting to spreadsheet '{file_path}' in sheet '{sheet_name}' with rules: {formatting_rules}.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False
            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]
                for cell in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for c in cell:
                        if 'font_name' in formatting_rules:
                            c.font = Font(name=formatting_rules['font_name'], size=formatting_rules.get('font_size', 11), bold=formatting_rules.get('bold', False), italic=formatting_rules.get('italic', False))
                        if 'alignment' in formatting_rules:
                            if formatting_rules['alignment'].lower() == 'center':
                                c.alignment = Alignment(horizontal='center')
                            elif formatting_rules['alignment'].lower() == 'left':
                                c.alignment = Alignment(horizontal='left')
                            elif formatting_rules['alignment'].lower() == 'right':
                                c.alignment = Alignment(horizontal='right')
                            elif formatting_rules['alignment'].lower() == 'justify':
                                c.alignment = Alignment(horizontal='justify')
                        if 'fill_color' in formatting_rules:
                            c.fill = PatternFill(start_color=formatting_rules['fill_color'], end_color=formatting_rules['fill_color'], fill_type='solid')
                wb.save(file_path)
            self.logger.info(f"Spreadsheet '{file_path}' formatted successfully in sheet '{sheet_name}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error formatting spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def add_chart(self, file_path: str, sheet_name: str, chart_type: str, data_range: str, chart_title: str, position: str = 'E5') -> bool:
        """
        Adds a chart to a specified sheet in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to add the chart.
            chart_type (str): The type of chart ('line', 'bar', 'pie', etc.).
            data_range (str): The Excel range containing the data (e.g., 'A1:B10').
            chart_title (str): The title of the chart.
            position (str, optional): The cell position to place the chart. Defaults to 'E5'.

        Returns:
            bool: True if the chart is added successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Adding {chart_type} chart to spreadsheet '{file_path}' in sheet '{sheet_name}' with data range '{data_range}' at position '{position}'.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False
            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]
                # Extract data
                data = ws[data_range]
                df = pd.DataFrame(data.values, columns=[cell.value for cell in data[0]])
                # Create chart using pandas
                chart = df.plot(kind=chart_type, title=chart_title)
                chart.figure.savefig('temp_chart.png')
                ws.add_image('temp_chart.png', position)
                wb.save(file_path)
                os.remove('temp_chart.png')
            self.logger.info(f"{chart_type.capitalize()} chart added successfully to '{file_path}' in sheet '{sheet_name}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error adding chart to spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def conditional_formatting(self, file_path: str, sheet_name: str, cell_range: str, rule: Dict[str, Any]) -> bool:
        """
        Applies conditional formatting to a specified range in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to apply formatting.
            cell_range (str): The Excel range to apply formatting (e.g., 'A1:A10').
            rule (Dict[str, Any]): A dictionary defining the conditional formatting rule.

        Returns:
            bool: True if conditional formatting is applied successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Applying conditional formatting to spreadsheet '{file_path}' in sheet '{sheet_name}' for range '{cell_range}' with rule: {rule}.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False
            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]
                from openpyxl.formatting.rule import CellIsRule
                from openpyxl.styles import Font, PatternFill

                if 'type' in rule and rule['type'] == 'cellIs':
                    operator = rule.get('operator', 'greaterThan')
                    formula = rule.get('formula', '0')
                    fill_color = rule.get('fill_color', 'FFFF00')
                    font_color = rule.get('font_color', '000000')
                    rule_obj = CellIsRule(operator=operator, formula=[formula],
                                          fill=PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
                                          font=Font(color=font_color))
                    ws.conditional_formatting.add(cell_range, rule_obj)
                    self.logger.debug(f"Conditional formatting rule applied to range '{cell_range}'.")
                else:
                    self.logger.error("Unsupported conditional formatting rule type.")
                    return False

                wb.save(file_path)
            self.logger.info(f"Conditional formatting applied successfully to '{file_path}' in sheet '{sheet_name}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error applying conditional formatting to spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def analyze_data(self, file_path: str, sheet_name: str, analysis_type: str, column: str) -> Optional[Any]:
        """
        Performs data analysis on a specified column in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name containing the data.
            analysis_type (str): The type of analysis ('mean', 'median', 'sum', 'max', 'min').
            column (str): The column letter to analyze (e.g., 'A').

        Returns:
            Optional[Any]: The result of the analysis, or None if analysis fails.
        """
        try:
            self.logger.debug(f"Analyzing data in spreadsheet '{file_path}' sheet '{sheet_name}' column '{column}' with analysis type '{analysis_type}'.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return None
            with self.lock:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                if column not in df.columns and column.upper() not in df.columns:
                    self.logger.error(f"Column '{column}' does not exist in sheet '{sheet_name}'.")
                    return None
                col_data = df[column].dropna()
                if analysis_type == 'mean':
                    result = col_data.mean()
                elif analysis_type == 'median':
                    result = col_data.median()
                elif analysis_type == 'sum':
                    result = col_data.sum()
                elif analysis_type == 'max':
                    result = col_data.max()
                elif analysis_type == 'min':
                    result = col_data.min()
                else:
                    self.logger.error(f"Unsupported analysis type '{analysis_type}'.")
                    return None
                self.logger.info(f"Analysis '{analysis_type}' on column '{column}' resulted in: {result}.")
                return result
        except Exception as e:
            self.logger.error(f"Error performing analysis on spreadsheet '{file_path}' sheet '{sheet_name}': {e}", exc_info=True)
            return None

    def add_pivot_table(self, file_path: str, sheet_name: str, pivot_sheet_name: str, data_range: str, index: List[str], columns: List[str], values: List[str], aggfunc: str = 'sum') -> bool:
        """
        Adds a pivot table to a specified sheet in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name containing the data.
            pivot_sheet_name (str): The sheet name to add the pivot table.
            data_range (str): The Excel range containing the data (e.g., 'A1:D100').
            index (List[str]): The list of columns to use as the pivot table index.
            columns (List[str]): The list of columns to use as the pivot table columns.
            values (List[str]): The list of columns to aggregate in the pivot table.
            aggfunc (str, optional): The aggregation function ('sum', 'mean', 'count', etc.). Defaults to 'sum'.

        Returns:
            bool: True if the pivot table is added successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Adding pivot table to spreadsheet '{file_path}' sheet '{pivot_sheet_name}' from data range '{data_range}'.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False
            with self.lock:
                df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=data_range)
                pivot_df = pd.pivot_table(df, index=index, columns=columns, values=values, aggfunc=aggfunc)
                pivot_wb = Workbook()
                pivot_ws = pivot_wb.active
                pivot_ws.title = pivot_sheet_name
                for r in dataframe_to_rows(pivot_df, index=True, header=True):
                    pivot_ws.append(r)
                pivot_wb.save('temp_pivot.xlsx')

                # Load the original workbook and add pivot sheet
                wb = load_workbook(file_path)
                temp_wb = load_workbook('temp_pivot.xlsx')
                temp_ws = temp_wb.active
                if pivot_sheet_name in wb.sheetnames:
                    wb.remove(wb[pivot_sheet_name])
                wb.add_sheet = temp_ws
                for row in temp_ws.iter_rows(values_only=True):
                    wb[pivot_sheet_name].append(row)
                wb.save(file_path)
                os.remove('temp_pivot.xlsx')
            self.logger.info(f"Pivot table added successfully to '{file_path}' in sheet '{pivot_sheet_name}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error adding pivot table to spreadsheet '{file_path}': {e}", exc_info=True)
            return False

    def add_data_validation(self, file_path: str, sheet_name: str, cell_range: str, validation_type: str, operator: Optional[str] = None, formula1: Optional[str] = None, formula2: Optional[str] = None, allow_blank: bool = True, show_input_message: bool = False, prompt: Optional[str] = None) -> bool:
        """
        Adds data validation to a specified range in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to apply data validation.
            cell_range (str): The Excel range to apply data validation (e.g., 'A1:A10').
            validation_type (str): The type of data validation ('whole', 'decimal', 'list', 'date', etc.).
            operator (Optional[str], optional): The operator for validation (e.g., 'greaterThan'). Defaults to None.
            formula1 (Optional[str], optional): The first formula or value. Defaults to None.
            formula2 (Optional[str], optional): The second formula or value for certain operators. Defaults to None.
            allow_blank (bool, optional): Whether to allow blank cells. Defaults to True.
            show_input_message (bool, optional): Whether to show an input message. Defaults to False.
            prompt (Optional[str], optional): The input message to display. Defaults to None.

        Returns:
            bool: True if data validation is added successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Adding data validation to spreadsheet '{file_path}' sheet '{sheet_name}' range '{cell_range}' with validation type '{validation_type}'.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False

            from openpyxl.worksheet.datavalidation import DataValidation

            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]
                dv = DataValidation(type=validation_type, operator=operator, formula1=formula1, formula2=formula2, allow_blank=allow_blank, showInputMessage=show_input_message, prompt=prompt)
                ws.add_data_validation(dv)
                dv.add(cell_range)
                wb.save(file_path)
            self.logger.info(f"Data validation added successfully to '{file_path}' in sheet '{sheet_name}' range '{cell_range}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error adding data validation to spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def apply_conditional_formatting(self, file_path: str, sheet_name: str, cell_range: str, rule: Dict[str, Any]) -> bool:
        """
        Applies conditional formatting to a specified range in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to apply conditional formatting.
            cell_range (str): The Excel range to apply formatting (e.g., 'A1:A10').
            rule (Dict[str, Any]): A dictionary defining the conditional formatting rule.

        Returns:
            bool: True if conditional formatting is applied successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Applying conditional formatting to spreadsheet '{file_path}' sheet '{sheet_name}' range '{cell_range}' with rule: {rule}.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False

            from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, IconSetRule
            from openpyxl.styles import Font, PatternFill, Color

            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]

                if rule['type'] == 'cellIs':
                    operator = rule.get('operator', 'greaterThan')
                    formula = rule.get('formula', '0')
                    fill_color = rule.get('fill_color', 'FFFF00')
                    font_color = rule.get('font_color', '000000')
                    rule_obj = CellIsRule(operator=operator, formula=[formula],
                                          fill=PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
                                          font=Font(color=font_color))
                    ws.conditional_formatting.add(cell_range, rule_obj)
                    self.logger.debug("CellIsRule conditional formatting applied.")
                elif rule['type'] == 'colorScale':
                    start_color = rule.get('start_color', 'FFFFFF')
                    mid_color = rule.get('mid_color', 'FF0000')
                    end_color = rule.get('end_color', '00FF00')
                    rule_obj = ColorScaleRule(start_type='num', start_value=0, start_color=start_color,
                                              mid_type='num', mid_value=50, mid_color=mid_color,
                                              end_type='num', end_value=100, end_color=end_color)
                    ws.conditional_formatting.add(cell_range, rule_obj)
                    self.logger.debug("ColorScaleRule conditional formatting applied.")
                elif rule['type'] == 'iconSet':
                    icons = rule.get('icons', '3TrafficLights1')
                    rule_obj = IconSetRule(icon_style=icons, showValue=True)
                    ws.conditional_formatting.add(cell_range, rule_obj)
                    self.logger.debug("IconSetRule conditional formatting applied.")
                else:
                    self.logger.error(f"Unsupported conditional formatting rule type '{rule['type']}'.")
                    return False

                wb.save(file_path)
            self.logger.info(f"Conditional formatting applied successfully to '{file_path}' in sheet '{sheet_name}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error applying conditional formatting to spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def create_pivot_table(self, file_path: str, source_sheet: str, pivot_sheet: str, pivot_table_name: str, data_range: str, rows: List[str], columns: List[str], values: List[str], aggfunc: str = 'sum') -> bool:
        """
        Creates a pivot table in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            source_sheet (str): The sheet name containing the source data.
            pivot_sheet (str): The sheet name to add the pivot table.
            pivot_table_name (str): The name of the pivot table.
            data_range (str): The Excel range containing the data (e.g., 'A1:D100').
            rows (List[str]): The list of columns to use as rows in the pivot table.
            columns (List[str]): The list of columns to use as columns in the pivot table.
            values (List[str]): The list of columns to aggregate in the pivot table.
            aggfunc (str, optional): The aggregation function ('sum', 'mean', 'count', etc.). Defaults to 'sum'.

        Returns:
            bool: True if the pivot table is created successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Creating pivot table '{pivot_table_name}' in spreadsheet '{file_path}' on sheet '{pivot_sheet}' from source sheet '{source_sheet}' range '{data_range}'.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False

            with self.lock:
                df = pd.read_excel(file_path, sheet_name=source_sheet, usecols=data_range)
                pivot_df = pd.pivot_table(df, index=rows, columns=columns, values=values, aggfunc=aggfunc)
                pivot_wb = Workbook()
                pivot_ws = pivot_wb.active
                pivot_ws.title = pivot_sheet
                for r in dataframe_to_rows(pivot_df, index=True, header=True):
                    pivot_ws.append(r)
                pivot_wb.save('temp_pivot.xlsx')

                # Load the original workbook and add pivot sheet
                wb = load_workbook(file_path)
                temp_wb = load_workbook('temp_pivot.xlsx')
                temp_ws = temp_wb.active
                if pivot_sheet in wb.sheetnames:
                    wb.remove(wb[pivot_sheet])
                pivot_ws_original = wb.create_sheet(title=pivot_sheet)
                for row in temp_ws.iter_rows(values_only=True):
                    pivot_ws_original.append(row)
                wb.save(file_path)
                os.remove('temp_pivot.xlsx')
            self.logger.info(f"Pivot table '{pivot_table_name}' created successfully in '{file_path}' on sheet '{pivot_sheet}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error creating pivot table '{pivot_table_name}' in spreadsheet '{file_path}': {e}", exc_info=True)
            return False

    def add_chart(self, file_path: str, sheet_name: str, chart_type: str, data_range: str, chart_title: str, position: str = 'E5') -> bool:
        """
        Adds a chart to a specified sheet in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to add the chart.
            chart_type (str): The type of chart ('line', 'bar', 'pie', etc.).
            data_range (str): The Excel range containing the data (e.g., 'A1:B10').
            chart_title (str): The title of the chart.
            position (str, optional): The cell position to place the chart. Defaults to 'E5'.

        Returns:
            bool: True if the chart is added successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Adding {chart_type} chart to spreadsheet '{file_path}' in sheet '{sheet_name}' with data range '{data_range}' at position '{position}'.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False

            from openpyxl.chart import LineChart, BarChart, PieChart, Reference

            chart_mapping = {
                'line': LineChart,
                'bar': BarChart,
                'pie': PieChart
            }

            if chart_type.lower() not in chart_mapping:
                self.logger.error(f"Unsupported chart type '{chart_type}'. Supported types: {list(chart_mapping.keys())}.")
                return False

            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]

                # Define the data for the chart
                data = Reference(ws, range_string=data_range)
                if chart_type.lower() == 'pie':
                    chart = chart_mapping[chart_type.lower()]()
                    categories = Reference(ws, min_col=data.min_col, min_row=data.min_row + 1, max_row=data.max_row)
                    chart.set_categories(categories)
                else:
                    chart = chart_mapping[chart_type.lower()]()
                chart.add_data(data, titles_from_data=True)
                chart.title = chart_title
                ws.add_chart(chart, position)
                wb.save(file_path)
            self.logger.info(f"{chart_type.capitalize()} chart added successfully to '{file_path}' in sheet '{sheet_name}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error adding {chart_type} chart to spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def add_formulas(self, file_path: str, sheet_name: str, formulas: Dict[str, str]) -> bool:
        """
        Adds formulas to specified cells in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to add formulas.
            formulas (Dict[str, str]): A dictionary where keys are cell addresses and values are formula strings.

        Returns:
            bool: True if formulas are added successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Adding formulas to spreadsheet '{file_path}' in sheet '{sheet_name}': {formulas}.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False

            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]
                for cell, formula in formulas.items():
                    ws[cell] = formula
                    self.logger.debug(f"Added formula '{formula}' to cell '{cell}'.")
                wb.save(file_path)
            self.logger.info(f"Formulas added successfully to '{file_path}' in sheet '{sheet_name}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error adding formulas to spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def conditional_formatting(self, file_path: str, sheet_name: str, cell_range: str, rule: Dict[str, Any]) -> bool:
        """
        Applies conditional formatting to a specified range in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to apply conditional formatting.
            cell_range (str): The Excel range to apply formatting (e.g., 'A1:A10').
            rule (Dict[str, Any]): A dictionary defining the conditional formatting rule.

        Returns:
            bool: True if conditional formatting is applied successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Applying conditional formatting to spreadsheet '{file_path}' sheet '{sheet_name}' range '{cell_range}' with rule: {rule}.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False

            from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, IconSetRule
            from openpyxl.styles import Font, PatternFill, Color

            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]

                if rule['type'] == 'cellIs':
                    operator = rule.get('operator', 'greaterThan')
                    formula = rule.get('formula', '0')
                    fill_color = rule.get('fill_color', 'FFFF00')
                    font_color = rule.get('font_color', '000000')
                    rule_obj = CellIsRule(operator=operator, formula=[formula],
                                          fill=PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid'),
                                          font=Font(color=font_color))
                    ws.conditional_formatting.add(cell_range, rule_obj)
                    self.logger.debug("CellIsRule conditional formatting applied.")
                elif rule['type'] == 'colorScale':
                    start_color = rule.get('start_color', 'FFFFFF')
                    mid_color = rule.get('mid_color', 'FF0000')
                    end_color = rule.get('end_color', '00FF00')
                    rule_obj = ColorScaleRule(start_type='num', start_value=0, start_color=start_color,
                                              mid_type='num', mid_value=50, mid_color=mid_color,
                                              end_type='num', end_value=100, end_color=end_color)
                    ws.conditional_formatting.add(cell_range, rule_obj)
                    self.logger.debug("ColorScaleRule conditional formatting applied.")
                elif rule['type'] == 'iconSet':
                    icons = rule.get('icons', '3TrafficLights1')
                    rule_obj = IconSetRule(icon_style=icons, showValue=True)
                    ws.conditional_formatting.add(cell_range, rule_obj)
                    self.logger.debug("IconSetRule conditional formatting applied.")
                else:
                    self.logger.error(f"Unsupported conditional formatting rule type '{rule['type']}'.")
                    return False

                wb.save(file_path)
            self.logger.info(f"Conditional formatting applied successfully to '{file_path}' in sheet '{sheet_name}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error applying conditional formatting to spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def add_data_validation(self, file_path: str, sheet_name: str, cell_range: str, validation_type: str, operator: Optional[str] = None, formula1: Optional[str] = None, formula2: Optional[str] = None, allow_blank: bool = True, show_input_message: bool = False, prompt: Optional[str] = None) -> bool:
        """
        Adds data validation to a specified range in an Excel spreadsheet.

        Args:
            file_path (str): The file path to the spreadsheet.
            sheet_name (str): The sheet name to apply data validation.
            cell_range (str): The Excel range to apply validation (e.g., 'A1:A10').
            validation_type (str): The type of data validation ('whole', 'decimal', 'list', 'date', etc.).
            operator (Optional[str], optional): The operator for validation (e.g., 'greaterThan'). Defaults to None.
            formula1 (Optional[str], optional): The first formula or value. Defaults to None.
            formula2 (Optional[str], optional): The second formula or value for certain operators. Defaults to None.
            allow_blank (bool, optional): Whether to allow blank cells. Defaults to True.
            show_input_message (bool, optional): Whether to show an input message. Defaults to False.
            prompt (Optional[str], optional): The input message to display. Defaults to None.

        Returns:
            bool: True if data validation is added successfully, False otherwise.
        """
        try:
            self.logger.debug(f"Adding data validation to spreadsheet '{file_path}' in sheet '{sheet_name}' range '{cell_range}' with validation type '{validation_type}'.")
            if not os.path.exists(file_path):
                self.logger.error(f"Spreadsheet '{file_path}' does not exist.")
                return False

            from openpyxl.worksheet.datavalidation import DataValidation

            with self.lock:
                wb = load_workbook(file_path)
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' does not exist in spreadsheet '{file_path}'.")
                    return False
                ws = wb[sheet_name]
                dv = DataValidation(type=validation_type, operator=operator, formula1=formula1, formula2=formula2, allow_blank=allow_blank, showInputMessage=show_input_message, prompt=prompt)
                ws.add_data_validation(dv)
                dv.add(cell_range)
                wb.save(file_path)
            self.logger.info(f"Data validation added successfully to '{file_path}' in sheet '{sheet_name}' range '{cell_range}'.")
            return True
        except Exception as e:
            self.logger.error(f"Error adding data validation to spreadsheet '{file_path}' in sheet '{sheet_name}': {e}", exc_info=True)
            return False

    def close_service(self):
        """
        Closes any resources or sessions held by the service.
        """
        try:
            self.logger.debug("Closing SpreadsheetService resources.")
            # Currently, no persistent resources to close
            self.logger.info("SpreadsheetService closed successfully.")
        except Exception as e:
            self.logger.error(f"Error closing SpreadsheetService: {e}", exc_info=True)
            raise SpreadsheetServiceError(f"Error closing SpreadsheetService: {e}")
